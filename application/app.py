from flask import Flask , request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from datetime import datetime
import xlsxwriter
import openpyxl
from openpyxl import Workbook

# file uploading
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from flask import send_from_directory



# Instantiate flask app
app = Flask(__name__)

# set configs
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:/// database.db"
app.config['SQLALCHEMY_TRACK_MODIFICATION'] = False


# Instentiate db object
db = SQLAlchemy(app) 

# Create marshmallow object
ma = Marshmallow(app)


# activate Workbook
wb = Workbook()
ws = wb.active

# Create database
class TodoList(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200),nullable=False)
    age = db.Column(db.Integer,nullable=False)
    email = db.Column(db.String(100),nullable=False) 
    phone = db.Column(db.Integer,nullable=False)
    date_created = db.Column(db.DateTime, default = datetime.utcnow)
    



    def __repr__(self) -> str:
        return f"{self.id}- {self.name}"


#Create TodoList Schema
class TodoListSchema(ma.Schema):
    class Meta:
        fields = ('id','name', 'age','email','phone','date_created')


#Create instance of schemas
todolist_schema = TodoListSchema(many=False)
todolists_schema = TodoListSchema(many=True)


@app.route("/fileparsing", methods =["POST"])
def add_single_todo():

    try:
        name = request.json['name']
        age =request.json['age']
        email =request.json['email']
        phone =request.json['phone']
    
        
        new_todo = TodoList(name = name , age = age ,email = email , phone = phone)
      

        db.session.add(new_todo)
        db.session.commit()

        return todolist_schema.jsonify(new_todo)

    except Exception as e:
        return jsonify({"Error":"Invalid request."})
    
# Add Multiple
@app.route("/fileparse", methods =["POST"])    
def addmultiple():
        data = request.get_json()
        print('data',data)
    
        for items in data:
            name = items["name"]
            age = items["age"]
            email = items["email"]
            phone = items["phone"]

            new_todo = TodoList(name=name , age=age ,email=email , phone=phone)
            db.session.add(new_todo)
            db.session.commit()
            new_todo = TodoList.query.all()

        return todolists_schema.jsonify(new_todo)
    
   


# Get todos
@app.route("/fileparsing", methods=["GET"])
def get_multiple_todo():
    todos = TodoList.query.all()
    result_set = todolists_schema.dump(todos)   
    return jsonify(result_set)




@app.route("/fileparsing/<int:id>", methods=["GET"])
def get_todo(id):
    todo = TodoList.query.get_or_404(int(id))
    return todolist_schema.jsonify(todo)

@app.route("/excelfile", methods=["GET"])
def alldata():
   
    dic = {}
    Alldata = TodoList.query.all()
    for i in Alldata:
        
        dic.update({"name":i.name, "age":i.age, "email": i.email, "phone":i.phone, "date_created":i.date_created})
        

    

    workbook = xlsxwriter.Workbook("python_toexcel.xlsx")
    worksheet = workbook.add_worksheet("firstSheet")

    worksheet.write('A1' ,"name")   
    worksheet.write('B1',"age") 
    worksheet.write('C1',"email")
    worksheet.write('D1',"phone")
    worksheet.write('E1',"date_created")
   

    for index , entry in enumerate(Alldata):
   
        worksheet.write(index+1, 0, entry.name)
        worksheet.write(index+1, 1, entry.age)
        worksheet.write(index+1, 2, entry.email)
        worksheet.write(index+1, 3, entry.phone)
        ws = entry.date_created.strftime('%x %X')
        # z= ws.strftime('%x %X')
        worksheet.write(index+1, 4, ws)
        


    workbook.close()
    return jsonify({"mesg":"Done"})


# Delete File Parsing
@app.route("/fileparsing/<int:id>", methods= ["DELETE"])
def delete_todo(id):
    todo = TodoList.query.get_or_404(int(id))
    db.session.delete(todo)
    db.session.commit()
    return jsonify({"Success": "Delete"})
     

# For upload file
@app.route("/uploadfile", methods= ["POST"])
def upload_file():
    file = request.files["filenames"]
    print("file..............", file)
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows, "rooooooooooooows")
    print(cols,"coluuuuuuuuuuuumn")

    for r in range(1,rows+1):
        for c in range(1, cols+1):
            print(sheet.cell(row=r, column=c).value, end = "     ")
        print()

    lst = []
    file.save(os.path.join('static', secure_filename(file.filename)))
    wb= load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):

        name = row[0].value
        age = row[1].value
        email = row[2].value
        phone = row[3].value
        # date_created =row[4].value
        lst.append([name,age, email,phone])
    
    for i in lst[1: ]:
        val = TodoList(name=i[0],age=i[1],email=i[2],phone=i[3])
        db.session.add(val)
    db.session.commit()
    return jsonify({"message":" Upload File successfully."})

# For Download file
def download_file(data):
    return send_from_directory('static',data, as_attachment =True)
    




if __name__ == "__main__":
    app.run(debug=True)


# create database
with app.app_context():
    db.create_all()